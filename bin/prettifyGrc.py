#!/usr/bin/env python3
"""
AmpRecon TSV to Excel converter.

Usage:
    python amprecon_convert.py input.tsv output.xlsx [--run-id RUN_ID]
    python amprecon_convert.py --help
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd


# ---------------------------------------------------------------------------
# Column mapping specification
# ---------------------------------------------------------------------------
# Each entry: (order_key, output_col, source_col_or_None, kind)
#   kind: 'rename'  - take from input, rename
#         'calc'    - compute from another column
#         'added'   - injected at runtime (GrcCreatedDate, RunId)
#         'drop'    - columns present in input that must be excluded
# ---------------------------------------------------------------------------

RENAME_MAP = {
    "ID":           "SampleId",
    "kelch13":      "Pfkelch13",
    "pm23-qpcr":    "pm23-Amp",
    "mdr1-qpcr":    "mdr1-Amp",
    "PGB:326":      "PfCRT:326",
    "PGB:356":      "PfCRT:356",
    "PGB:127":      "PfARPS10:127",
    "PGB:128":      "PfARPS10:128",
    "PGB:193":      "PfFD:193",
    "PGB:484":      "PfMDR2:484",
    "McCOIL":       "COI",
    "Barcode":      "GenBarcode",
}

# Columns that should be completely dropped from the output
DROP_COLS = {"sample_id", "primer_panel", "PfEXO", "P23:BP"}

# Final ordered column list (matches the Order column in AmpRecon sheet)
ORDERED_COLS = [
    # order 1R
    "SampleId",
    # order 153-161, 158, 159 — drug resistance phenotypes
    "Artemisinin",
    "Piperaquine",
    "Mefloquine",
    "Chloroquine",
    "Pyrimethamine",
    "Sulfadoxine",
    "DHA-PPQ",
    "AS-MQ",
    "S-P",
    "S-P-IPTp",
    # order 2
    "Species",
    # order 5R
    "GenBarcode",
    # New calculated
    "GenBarcodeMissing",
    "GenBarcodeHet",
    # Added (COI / McCOIL)
    "COI",
    # order 3R
    "Pfkelch13",
    # order 6-11
    "PfCRT",
    "PfDHFR",
    "PfDHPS",
    "PfMDR1",
    # order 150R, 149R
    "pm23-Amp",
    "mdr1-Amp",
    # order 11
    "PGB",
    # PfCRT positions 12-24 + renamed
    "PfCRT:72", "PfCRT:74", "PfCRT:75", "PfCRT:76", "PfCRT:93", "PfCRT:97",
    "PfCRT:218", "PfCRT:220", "PfCRT:271",
    "PfCRT:326",   # renamed from PGB:326
    "PfCRT:333", "PfCRT:353",
    "PfCRT:356",   # renamed from PGB:356
    "PfCRT:371",
    # PfDHFR positions
    "PfDHFR:16", "PfDHFR:51", "PfDHFR:59", "PfDHFR:108", "PfDHFR:164", "PfDHFR:306",
    # PfDHPS positions
    "PfDHPS:436", "PfDHPS:437", "PfDHPS:540", "PfDHPS:581", "PfDHPS:613",
    # PfEXO:415
    "PfEXO:415",
    # PfMDR1 positions
    "PfMDR1:86", "PfMDR1:184", "PfMDR1:1034", "PfMDR1:1042", "PfMDR1:1226", "PfMDR1:1246",
    # Renamed PGB positions
    "PfARPS10:127", "PfARPS10:128", "PfFD:193", "PfMDR2:484",
    # Barcode SNP columns 48-148
    "Pf3D7_02_v3:376222", "Pf3D7_02_v3:470013", "Pf3D7_03_v3:656861",
    "Pf3D7_04_v3:110442", "Pf3D7_04_v3:881571", "Pf3D7_05_v3:350933",
    "Pf3D7_05_v3:369740", "Pf3D7_06_v3:900278", "Pf3D7_07_v3:1044052",
    "Pf3D7_08_v3:1314831", "Pf3D7_08_v3:413067", "Pf3D7_09_v3:900277",
    "Pf3D7_11_v3:1018899", "Pf3D7_11_v3:1815412", "Pf3D7_13_v3:1056452",
    "Pf3D7_13_v3:1466422", "Pf3D7_14_v3:137622", "Pf3D7_14_v3:2164225",
    "Pf3D7_01_v3:145515", "Pf3D7_03_v3:548178", "Pf3D7_04_v3:1102392",
    "Pf3D7_04_v3:139051", "Pf3D7_04_v3:286542", "Pf3D7_04_v3:529500",
    "Pf3D7_05_v3:796714", "Pf3D7_07_v3:1256331", "Pf3D7_07_v3:461139",
    "Pf3D7_07_v3:619957", "Pf3D7_08_v3:417335", "Pf3D7_09_v3:163977",
    "Pf3D7_10_v3:317581", "Pf3D7_10_v3:336274", "Pf3D7_11_v3:1020397",
    "Pf3D7_11_v3:1294107", "Pf3D7_11_v3:1935227", "Pf3D7_11_v3:477922",
    "Pf3D7_12_v3:1663492", "Pf3D7_12_v3:2171901", "Pf3D7_13_v3:1233218",
    "Pf3D7_13_v3:1867630", "Pf3D7_13_v3:2377887", "Pf3D7_14_v3:2355751",
    "Pf3D7_14_v3:3046108", "Pf3D7_02_v3:529709", "Pf3D7_02_v3:714480",
    "Pf3D7_03_v3:155697", "Pf3D7_04_v3:1037656", "Pf3D7_04_v3:648101",
    "Pf3D7_05_v3:1204155", "Pf3D7_06_v3:1282691", "Pf3D7_06_v3:1289212",
    "Pf3D7_07_v3:1066698", "Pf3D7_07_v3:1213486", "Pf3D7_07_v3:704373",
    "Pf3D7_08_v3:1313202", "Pf3D7_08_v3:339406", "Pf3D7_08_v3:701557",
    "Pf3D7_09_v3:452690", "Pf3D7_09_v3:599655", "Pf3D7_10_v3:1383789",
    "Pf3D7_10_v3:1385894", "Pf3D7_11_v3:1006911", "Pf3D7_11_v3:1295068",
    "Pf3D7_11_v3:1802201", "Pf3D7_12_v3:1667593", "Pf3D7_12_v3:1934745",
    "Pf3D7_12_v3:858501", "Pf3D7_13_v3:1419519", "Pf3D7_13_v3:159086",
    "Pf3D7_13_v3:2161975", "Pf3D7_13_v3:2573828", "Pf3D7_13_v3:388365",
    "Pf3D7_14_v3:2625887", "Pf3D7_14_v3:3126219", "Pf3D7_14_v3:438592",
    "Pf3D7_01_v3:179347", "Pf3D7_01_v3:180554", "Pf3D7_01_v3:283144",
    "Pf3D7_01_v3:535211", "Pf3D7_02_v3:839620", "Pf3D7_04_v3:426436",
    "Pf3D7_04_v3:531138", "Pf3D7_04_v3:891732", "Pf3D7_05_v3:172801",
    "Pf3D7_06_v3:574938", "Pf3D7_07_v3:1308383", "Pf3D7_07_v3:1358910",
    "Pf3D7_07_v3:1359218", "Pf3D7_07_v3:635985", "Pf3D7_08_v3:1056829",
    "Pf3D7_08_v3:150033", "Pf3D7_08_v3:399774", "Pf3D7_09_v3:1379145",
    "Pf3D7_10_v3:1386850", "Pf3D7_11_v3:1935031", "Pf3D7_11_v3:408668",
    "Pf3D7_11_v3:828596", "Pf3D7_12_v3:857245", "Pf3D7_14_v3:107014",
    "Pf3D7_14_v3:1757603", "Pf3D7_14_v3:2733656",
    # Added metadata
    "GrcCreatedDate",
    "RunId",
    # Raw paths (order 164, 165)
    "fastq1_path",
    "fastq2_path",
    # order 151
    "species-qpcr",
]

# Drug phenotype columns (present in pipeline output, listed here to ensure ordering)
DRUG_PHENOTYPE_COLS = [
    "Artemisinin", "Piperaquine", "Mefloquine", "Chloroquine",
    "Pyrimethamine", "Sulfadoxine", "DHA-PPQ", "AS-MQ", "S-P",
]

# ---------------------------------------------------------------------------
# Header colour zones — exact RGB values from ExampleFormat.xlsx
# Each tuple: (fill_hex, [list of column names])
# ---------------------------------------------------------------------------
# Grey  — sample identity / administrative grouping
COLOUR_GREY    = "BFBFBF"
# Pink  — site / collection metadata
COLOUR_PINK    = "EEA9B8"
# Gold  — drug resistance phenotypes
COLOUR_GOLD    = "EEB422"
# Cyan  — barcode / species
COLOUR_CYAN    = "8EE5EE"
# Green — summary drug markers
COLOUR_GREEN   = "7CCD7C"
# Blue  — individual SNP positions
COLOUR_BLUE    = "5CACEE"
# Purple — qPCR / amplicon ancillary
COLOUR_PURPLE  = "AB82FF"
# Red   — barcode SNP columns
COLOUR_RED     = "EE6363"

# Map each output column to its header colour
COL_COLOURS: dict[str, str] = {}

def _assign(cols, colour):
    for c in cols:
        COL_COLOURS[c] = colour

_assign(["SampleId", "SeqNum", "Source", "Process", "Study", "Year", "Month",
         "TimePoint", "Country", "AdmDiv1", "AdmDiv1_GID", "AdmDiv2", "AdmDiv2_GID"],
        COLOUR_GREY)

_assign(["SiteName", "Latitude", "Longitude", "CollectionDate",
         "CollectionID", "ManifestID", "DataRelease"],
        COLOUR_PINK)

_assign(["Artemisinin", "Piperaquine", "Mefloquine", "Chloroquine",
         "Pyrimethamine", "Sulfadoxine", "DHA-PPQ", "AS-MQ", "S-P", "S-P-IPTp"],
        COLOUR_GOLD)

_assign(["Species", "GenBarcode", "GenBarcodeMissing", "GenBarcodeHet", "COI"],
        COLOUR_CYAN)

_assign(["Pfkelch13", "PfCRT", "PfDHFR", "PfDHPS", "PfMDR1",
         "pm23-Amp", "mdr1-Amp", "PGB"],
        COLOUR_GREEN)

_assign(["PfCRT:72", "PfCRT:74", "PfCRT:75", "PfCRT:76", "PfCRT:93", "PfCRT:97",
         "PfCRT:218", "PfCRT:220", "PfCRT:271", "PfCRT:326", "PfCRT:333", "PfCRT:353",
         "PfCRT:356", "PfCRT:371",
         "PfDHFR:16", "PfDHFR:51", "PfDHFR:59", "PfDHFR:108", "PfDHFR:164", "PfDHFR:306",
         "PfDHPS:436", "PfDHPS:437", "PfDHPS:540", "PfDHPS:581", "PfDHPS:613",
         "PfEXO:415",
         "PfMDR1:86", "PfMDR1:184", "PfMDR1:1034", "PfMDR1:1042", "PfMDR1:1226", "PfMDR1:1246",
         "PfARPS10:127", "PfARPS10:128", "PfFD:193", "PfMDR2:484"],
        COLOUR_BLUE)

_assign(["pm23-break", "pm23-qPCR", "mdr1-qPCR",
         "species-aSeq", "species-qPCR", "species-barcode", "species-qpcr"],
        COLOUR_PURPLE)

# All Pf3D7 barcode SNP columns → red (matched by prefix in apply_header_colours)


def calc_barcode_missing(barcode_series: pd.Series) -> pd.Series:
    """Ratio of 'X' characters in GenBarcode string."""
    def ratio(val):
        if pd.isna(val) or not isinstance(val, str) or len(val) == 0:
            return pd.NA
        return val.count("X") / len(val)
    return barcode_series.apply(ratio)


def calc_barcode_het(barcode_series: pd.Series) -> pd.Series:
    """Ratio of 'N' characters in GenBarcode string."""
    def ratio(val):
        if pd.isna(val) or not isinstance(val, str) or len(val) == 0:
            return pd.NA
        return val.count("N") / len(val)
    return barcode_series.apply(ratio)


def load_metadata(meta_path: str) -> pd.DataFrame:
    """Load optional metadata file (TSV or Excel)."""
    p = Path(meta_path)
    if p.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(p, dtype=str)
    return pd.read_csv(p, sep="\t", dtype=str, keep_default_na=False, na_values=[""])


def validate_and_join_metadata(df_main: pd.DataFrame, df_meta: pd.DataFrame) -> pd.DataFrame:
    """
    Validate that metadata matches main data on row count and SampleId,
    report any discrepancies, then left-join metadata columns into df_main
    (inserted after SampleId, before all other columns).
    """
    issues = []

    # Row count check
    if len(df_meta) != len(df_main):
        issues.append(
            f"  [MISMATCH] Row count: pipeline has {len(df_main)} rows, "
            f"metadata has {len(df_meta)} rows."
        )

    if "SampleId" not in df_meta.columns:
        issues.append("  [ERROR] Metadata file has no 'SampleId' column — cannot join.")
        if issues:
            print("\n".join(issues))
        return df_main

    # SampleId set comparison
    main_ids  = set(df_main["SampleId"].astype(str))
    meta_ids  = set(df_meta["SampleId"].astype(str))
    only_main = sorted(main_ids - meta_ids)
    only_meta = sorted(meta_ids - main_ids)
    if only_main:
        issues.append(f"  [MISMATCH] SampleIds in pipeline but NOT in metadata ({len(only_main)}): {only_main[:10]}{'...' if len(only_main) > 10 else ''}")
    if only_meta:
        issues.append(f"  [MISMATCH] SampleIds in metadata but NOT in pipeline ({len(only_meta)}): {only_meta[:10]}{'...' if len(only_meta) > 10 else ''}")

    if issues:
        print("Metadata validation warnings:")
        for msg in issues:
            print(msg)
    else:
        print(f"  Metadata validated OK: {len(df_meta)} rows, SampleIds match.")

    # Join: metadata cols (excluding SampleId) inserted right after SampleId in main df
    meta_extra_cols = [c for c in df_meta.columns if c != "SampleId"]
    df_merged = df_main.merge(df_meta[["SampleId"] + meta_extra_cols],
                               on="SampleId", how="left", suffixes=("", "_meta"))

    # Reorder: SampleId, then meta cols, then rest
    rest = [c for c in df_merged.columns if c != "SampleId" and c not in meta_extra_cols]
    df_merged = df_merged[["SampleId"] + meta_extra_cols + rest]
    return df_merged


def apply_header_colours(ws, col_names: list[str]) -> None:
    """Apply per-column background colours to the header row of a worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(bold=True, color="FF000000", name="Calibri", size=10)
    no_border = Border(
        left=Side(border_style=None),
        right=Side(border_style=None),
        top=Side(border_style=None),
        bottom=Side(border_style=None),
    )

    for idx, col_name in enumerate(col_names, start=1):
        cell = ws.cell(row=1, column=idx)
        # Determine colour
        if col_name.startswith("Pf3D7_"):
            hex_colour = COLOUR_RED
        else:
            hex_colour = COL_COLOURS.get(col_name, "BFBFBF")
        cell.fill = PatternFill("solid", start_color=hex_colour)
        cell.font = header_font
        cell.border = no_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_sheet(ws, col_names: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    # Auto column widths
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

    apply_header_colours(ws, col_names)
    ws.freeze_panes = "A2"


def convert(tsv_path: str, xlsx_path: str, run_id: str = "",
            meta_path: str | None = None) -> None:
    print(f"Reading: {tsv_path}")
    df = pd.read_csv(tsv_path, sep="\t", dtype=str, keep_default_na=False, na_values=[""])
    print(f"  {len(df)} rows, {len(df.columns)} columns found in input.")

    # 1. Rename columns
    df = df.rename(columns=RENAME_MAP)

    # 2. Drop unwanted columns
    cols_to_drop = [c for c in df.columns if c in DROP_COLS]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
        print(f"  Dropped columns: {cols_to_drop}")

    # 3. Calculate GenBarcodeMissing and GenBarcodeHet from GenBarcode
    if "GenBarcode" in df.columns:
        df["GenBarcodeMissing"] = calc_barcode_missing(df["GenBarcode"])
        df["GenBarcodeHet"] = calc_barcode_het(df["GenBarcode"])
    else:
        df["GenBarcodeMissing"] = pd.NA
        df["GenBarcodeHet"] = pd.NA
        print("  Warning: 'GenBarcode' column not found; GenBarcodeMissing/Het set to NA.")

    # 4. COI / McCOIL fallback
    if "COI" not in df.columns:
        print("  COI (McCOIL) column not present — filling with 'NA'.")
        df["COI"] = "NA"

    # 5. Add runtime metadata columns
    df["GrcCreatedDate"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df["RunId"] = run_id if run_id else ""

    # 6. Optional metadata join (grey columns)
    if meta_path:
        print(f"Loading metadata: {meta_path}")
        df_meta = load_metadata(meta_path)
        df = validate_and_join_metadata(df, df_meta)

    # 7. Reorder columns
    runinfo_cols = ["RunId", "GrcCreatedDate", "fastq1_path", "fastq2_path"]
    # Metadata extra cols go right after SampleId
    meta_extra_cols = []
    if meta_path:
        try:
            _df_meta_check = load_metadata(meta_path)
            meta_extra_cols = [c for c in _df_meta_check.columns
                               if c != "SampleId" and c in df.columns
                               and c not in set(ORDERED_COLS) and c not in runinfo_cols]
        except Exception:
            pass
    known = set(ORDERED_COLS) | set(runinfo_cols) | set(meta_extra_cols)
    ampRecon_ordered = [c for c in ORDERED_COLS if c in df.columns and c not in runinfo_cols]
    if meta_extra_cols and "SampleId" in ampRecon_ordered:
        idx = ampRecon_ordered.index("SampleId") + 1
        ampRecon_ordered = ampRecon_ordered[:idx] + meta_extra_cols + ampRecon_ordered[idx:]
    extra_cols = [c for c in df.columns if c not in known and c not in runinfo_cols]
    if extra_cols:
        print(f"  Extra columns not in spec (appended at end): {extra_cols}")
    ampRecon_cols = ampRecon_ordered + extra_cols
    df_amprecon = df[ampRecon_cols]

    # Build RunInfo sheet
    runinfo_present = ["SampleId"] + [c for c in runinfo_cols if c in df.columns]
    df_runinfo = df[runinfo_present]

    # 8. Write to Excel
    print(f"Writing: {xlsx_path}")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_amprecon.to_excel(writer, sheet_name="AmpRecon", index=False)
        df_runinfo.to_excel(writer, sheet_name="RunInfo", index=False)
        style_sheet(writer.sheets["AmpRecon"], ampRecon_cols)
        style_sheet(writer.sheets["RunInfo"], runinfo_present)

    print(f"Done. Output saved to: {xlsx_path}")
    print(f"  AmpRecon columns: {len(ampRecon_cols)}")
    print(f"  RunInfo columns:  {len(runinfo_present)}")


def main():
    parser = argparse.ArgumentParser(
        description="Convert AmpRecon tab-delimited pipeline output to Excel (.xlsx).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python amprecon_convert.py results.tsv output.xlsx
  python amprecon_convert.py results.tsv output.xlsx --run-id RUN_2024_001
  python amprecon_convert.py results.tsv output.xlsx --metadata meta.tsv
  python amprecon_convert.py results.tsv output.xlsx --metadata meta.xlsx --run-id RUN_001
        """,
    )
    parser.add_argument("input",  help="Path to input tab-delimited (.tsv / .txt) file")
    parser.add_argument("output", help="Path for output Excel (.xlsx) file")
    parser.add_argument(
        "--run-id", default="", metavar="RUN_ID",
        help="User-specified run identifier written into the RunId column (default: empty)",
    )
    parser.add_argument(
        "--metadata", default=None, metavar="FILE",
        help=(
            "Optional metadata file (.tsv or .xlsx) whose columns are inserted as grey "
            "columns after SampleId. Must contain a 'SampleId' column. "
            "Row count and SampleIds are validated against the pipeline output; "
            "any mismatches are reported."
        ),
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    if args.metadata and not Path(args.metadata).exists():
        print(f"Error: metadata file not found: {args.metadata}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        convert(str(input_path), str(output_path),
                run_id=args.run_id, meta_path=args.metadata)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        import traceback; traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
